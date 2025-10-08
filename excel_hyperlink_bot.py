import openpyxl
from openpyxl.styles import Font
import re
import os
import shutil
from datetime import datetime
import json
import logging

# Try to import tqdm for progress bars, with fallback
try:
    from tqdm import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False
    print("⚠️  tqdm not installed. Progress bars disabled. Install with: pip install tqdm")

# ==================== CONFIGURATION SYSTEM ====================
class BotConfig:
    def __init__(self):
        self.config_file = 'excel_bot_config.json'
        self.defaults = {
            "hyperlink_color": "0000FF",
            "backup_files": True,
            "auto_detect": True,
            "max_rows_to_process": 100000,
            "supported_extensions": [".xlsx", ".xlsm", ".xltx", ".xltm"],
            "social_media_platforms": ["linkedin", "twitter", "facebook", "instagram", "youtube"],
            "log_level": "INFO"
        }
        self.load_config()
    
    def load_config(self):
        """Load configuration from file or create default"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r') as f:
                    saved_config = json.load(f)
                    self.defaults.update(saved_config)
                    logging.info("Configuration loaded from file")
            else:
                self.save_config()
                logging.info("Default configuration created")
        except Exception as e:
            logging.warning(f"Could not load config: {e}. Using defaults.")
    
    def save_config(self):
        """Save current configuration to file"""
        try:
            with open(self.config_file, 'w') as f:
                json.dump(self.defaults, f, indent=2)
            logging.info("Configuration saved")
        except Exception as e:
            logging.error(f"Could not save config: {e}")
    
    def get(self, key):
        return self.defaults.get(key)
    
    def set(self, key, value):
        self.defaults[key] = value
        self.save_config()

# Initialize configuration
config = BotConfig()

# ==================== LOGGING SETUP ====================
def setup_logging():
    """Setup logging with rotation"""
    log_file = 'excel_bot.log'
    
    # Clear log if it gets too large (>10MB)
    if os.path.exists(log_file) and os.path.getsize(log_file) > 10 * 1024 * 1024:
        os.remove(log_file)
    
    logging.basicConfig(
        level=getattr(logging, config.get('log_level')),
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()  # Also print to console
        ]
    )
    logging.info("=== Excel Hyperlink Bot Started ===")

# ==================== BACKUP SYSTEM ====================
def create_backup(file_path):
    """Create timestamped backup before conversion"""
    if not config.get('backup_files'):
        return None
    
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_dir = "backups"
        
        # Create backups directory if it doesn't exist
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        file_name = os.path.basename(file_path)
        backup_path = os.path.join(backup_dir, f"{file_name}.backup_{timestamp}")
        
        shutil.copy2(file_path, backup_path)
        logging.info(f"Backup created: {backup_path}")
        return backup_path
    
    except Exception as e:
        logging.error(f"Backup failed: {e}")
        return None

def restore_backup(backup_path, original_path):
    """Restore from backup if needed"""
    try:
        shutil.copy2(backup_path, original_path)
        logging.info(f"Restored from backup: {backup_path}")
        return True
    except Exception as e:
        logging.error(f"Restore failed: {e}")
        return False

# ==================== CORE DETECTION FUNCTIONS ====================
def is_valid_email(text):
    """Check if text is a valid email address"""
    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(email_pattern, str(text).strip()) is not None

def is_valid_website(text):
    """Check if text is a valid website URL"""
    text = str(text).strip().lower()
    if not text or '.' not in text:
        return False
    
    # Common domain extensions
    common_domains = [
        '.com', '.org', '.net', '.edu', '.gov', '.io', '.co', '.info', '.in',
        '.us', '.uk', '.ca', '.au', '.de', '.fr', '.jp', '.cn', '.br', '.ru',
        '.mx', '.es', '.it', '.nl', '.se', '.no', '.dk', '.fi', '.pl', '.ch'
    ]
    
    # Website pattern
    website_pattern = r'^(https?://)?(www\.)?[a-zA-Z0-9-]+\.[a-zA-Z]{2,}(?:\.[a-zA-Z]{2,})?(?:/\S*)?$'
    
    return (re.match(website_pattern, text) is not None) or any(ext in text for ext in common_domains)

def is_valid_linkedin(text):
    """Check if text is a LinkedIn URL"""
    text = str(text).strip().lower()
    return 'linkedin.com' in text

def format_website_url(url):
    """Format website URL by adding https:// if missing"""
    url = str(url).strip()
    if not url.startswith(('http://', 'https://')):
        return 'https://' + url
    return url

def detect_content_type(text):
    """Detect what type of content the text contains"""
    text = str(text).strip()
    
    if not text:
        return "empty"
    
    if is_valid_email(text):
        return "email"
    elif is_valid_linkedin(text):
        return "linkedin"
    elif is_valid_website(text):
        return "website"
    else:
        return "text"

def analyze_column(ws, column_letter):
    """Analyze a column to detect what type of content it contains"""
    content_types = {}
    sample_size = min(100, ws.max_row)
    
    for row in range(1, sample_size + 1):
        cell = ws[f'{column_letter}{row}']
        value = str(cell.value).strip() if cell.value else ""
        
        if value:
            content_type = detect_content_type(value)
            content_types[content_type] = content_types.get(content_type, 0) + 1
    
    if content_types:
        return max(content_types.items(), key=lambda x: x[1])
    return ("unknown", 0)

# ==================== PROGRESS-BASED CONVERSION ====================
def convert_column_to_hyperlinks(ws, column_letter, column_name):
    """Convert a specific column to hyperlinks with progress tracking"""
    hyperlink_font = Font(color=config.get('hyperlink_color'), underline='single')
    converted_count = 0
    
    # Analyze column content
    content_type, confidence = analyze_column(ws, column_letter)
    print(f"📊 Column {column_letter} ({column_name}): {content_type} (confidence: {confidence})")
    
    total_rows = min(ws.max_row, config.get('max_rows_to_process'))
    
    # Use progress bar if available
    if TQDM_AVAILABLE:
        progress_bar = tqdm(
            total=total_rows, 
            desc=f"🔄 {column_letter}", 
            leave=False,
            bar_format='{l_bar}{bar:30}{r_bar}',
            ncols=80
        )
    else:
        print(f"⏳ Processing {column_letter} ({total_rows} rows)...")
        last_progress = 0
    
    for row in range(1, total_rows + 1):
        cell = ws[f'{column_letter}{row}']
        value = str(cell.value).strip() if cell.value else ""
        
        if value:
            content_type = detect_content_type(value)
            
            if content_type == "email":
                cell.hyperlink = f'mailto:{value}'
                cell.font = hyperlink_font
                converted_count += 1
                
            elif content_type in ["website", "linkedin"]:
                formatted_url = format_website_url(value)
                cell.hyperlink = formatted_url
                cell.font = hyperlink_font
                converted_count += 1
        
        # Update progress
        if TQDM_AVAILABLE:
            progress_bar.update(1)
        else:
            # Simple progress indicator without tqdm
            progress = (row * 100) // total_rows
            if progress >= last_progress + 10:  # Update every 10%
                print(f"   {progress}% complete ({row}/{total_rows} rows)")
                last_progress = progress
    
    if TQDM_AVAILABLE:
        progress_bar.close()
    
    return converted_count

# ==================== BATCH PROCESSING ====================
def batch_process_folder(folder_path):
    """Process all Excel files in a folder"""
    if not os.path.exists(folder_path):
        logging.error(f"Folder not found: {folder_path}")
        print("❌ Folder not found!")
        return False
    
    excel_files = []
    for file in os.listdir(folder_path):
        if any(file.lower().endswith(ext) for ext in config.get('supported_extensions')):
            excel_files.append(os.path.join(folder_path, file))
    
    if not excel_files:
        logging.warning(f"No Excel files found in: {folder_path}")
        print("❌ No Excel files found in the folder!")
        return False
    
    print(f"📁 Found {len(excel_files)} Excel files to process")
    logging.info(f"Found {len(excel_files)} Excel files to process")
    
    success_count = 0
    
    # Use progress bar for batch processing if available
    if TQDM_AVAILABLE:
        batch_progress = tqdm(excel_files, desc="Processing folder", leave=True)
    else:
        batch_progress = excel_files
        print("🔄 Starting batch processing...")
    
    for i, file_path in enumerate(batch_progress):
        try:
            if process_single_file(file_path):
                success_count += 1
            else:
                logging.error(f"Failed to process: {file_path}")
        except Exception as e:
            logging.error(f"Error processing {file_path}: {e}")
            print(f"❌ Error processing {os.path.basename(file_path)}: {e}")
        
        # Update progress for non-tqdm
        if not TQDM_AVAILABLE:
            progress = ((i + 1) * 100) // len(excel_files)
            print(f"📦 Batch progress: {progress}% ({i + 1}/{len(excel_files)} files)")
    
    if TQDM_AVAILABLE:
        batch_progress.close()
    
    print(f"\n✅ Batch processing completed: {success_count}/{len(excel_files)} files successful")
    logging.info(f"Batch processing completed: {success_count}/{len(excel_files)} files successful")
    return success_count > 0

# ==================== SINGLE FILE PROCESSING ====================
def process_single_file(file_path):
    """Process a single Excel file with all features"""
    try:
        logging.info(f"Starting processing: {file_path}")
        
        # Create backup
        backup_path = create_backup(file_path)
        
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        print(f"\n📁 Processing: {os.path.basename(file_path)}")
        print(f"📏 Dimensions: {ws.max_row} rows × {ws.max_column} columns")
        
        total_converted = 0
        
        # Process each column with smart detection
        for col in range(1, ws.max_column + 1):
            column_letter = openpyxl.utils.get_column_letter(col)
            header = ws.cell(row=1, column=col).value
            column_name = header if header else f"Column {column_letter}"
            
            converted = convert_column_to_hyperlinks(ws, column_letter, column_name)
            total_converted += converted
        
        # Save the workbook
        base_name, ext = os.path.splitext(file_path)
        output_path = f"{base_name}_with_hyperlinks{ext}"
        wb.save(output_path)
        
        logging.info(f"Conversion completed: {total_converted} links converted")
        print(f"\n✅ Conversion completed!")
        print(f"🔗 Total hyperlinks created: {total_converted}")
        print(f"💾 Saved as: {output_path}")
        if backup_path:
            print(f"📂 Backup: {backup_path}")
        
        return True
        
    except Exception as e:
        logging.error(f"Error processing {file_path}: {e}")
        print(f"❌ Error: {str(e)}")
        return False

# ==================== CONFIGURATION MENUS ====================
def show_configuration():
    """Show current configuration"""
    print("\n⚙️  Current Configuration:")
    print("-" * 40)
    for key, value in config.defaults.items():
        print(f"  {key}: {value}")
    
    print(f"\n📊 System Info:")
    print(f"  Progress bars: {'Enabled' if TQDM_AVAILABLE else 'Disabled'}")
    print(f"  Python version: {os.sys.version.split()[0]}")

def update_configuration():
    """Allow user to update configuration"""
    while True:
        print("\n📝 Update Configuration:")
        print("1. Change hyperlink color (current: {})".format(config.get('hyperlink_color')))
        print("2. Toggle backup files (current: {})".format(config.get('backup_files')))
        print("3. Set max rows to process (current: {})".format(config.get('max_rows_to_process')))
        print("4. Change log level (current: {})".format(config.get('log_level')))
        print("5. Back to main menu")
        
        choice = input("\nEnter your choice (1-5): ").strip()
        
        if choice == "1":
            new_color = input("Enter hex color (e.g., 0000FF for blue): ").strip()
            if re.match(r'^[0-9A-Fa-f]{6}$', new_color):
                config.set("hyperlink_color", new_color.upper())
                print("✅ Color updated!")
            else:
                print("❌ Invalid color format. Use 6 hex characters (e.g., 0000FF)")
        
        elif choice == "2":
            current = config.get("backup_files")
            new_value = not current
            config.set("backup_files", new_value)
            print(f"✅ Backup files: {new_value}")
        
        elif choice == "3":
            try:
                max_rows = int(input("Enter max rows to process: ").strip())
                if max_rows > 0:
                    config.set("max_rows_to_process", max_rows)
                    print("✅ Max rows updated!")
                else:
                    print("❌ Please enter a positive number")
            except ValueError:
                print("❌ Please enter a valid number")
        
        elif choice == "4":
            levels = ["DEBUG", "INFO", "WARNING", "ERROR"]
            print("Available levels: " + ", ".join(levels))
            new_level = input("Enter log level: ").strip().upper()
            if new_level in levels:
                config.set("log_level", new_level)
                setup_logging()  # Re-initialize logging
                print("✅ Log level updated!")
            else:
                print("❌ Invalid log level")
        
        elif choice == "5":
            break
        
        else:
            print("❌ Invalid choice!")

def view_log_file():
    """Display the log file contents"""
    log_file = 'excel_bot.log'
    if os.path.exists(log_file):
        print(f"\n📋 Last 20 log entries from {log_file}:")
        print("=" * 60)
        try:
            with open(log_file, 'r', encoding='utf-8') as f:
                lines = f.readlines()[-20:]
                for line in lines:
                    print(line.strip())
        except Exception as e:
            print(f"❌ Could not read log file: {e}")
    else:
        print("❌ Log file not found yet. Process some files first.")

# ==================== MAIN MENU ====================
def main():
    setup_logging()
    
    print("🚀 ENHANCED Excel Hyperlink Converter Bot")
    print("==========================================")
    print("✨ Features: Backup System • Batch Processing • Progress Bars • Logging • Configuration")
    
    if not TQDM_AVAILABLE:
        print("\n⚠️  For progress bars, install: pip install tqdm")
    
    while True:
        print("\n🎯 Main Menu:")
        print("1. Process single Excel file")
        print("2. Process folder (batch mode)")
        print("3. View configuration")
        print("4. Update configuration")
        print("5. View log file")
        print("6. Exit")
        
        choice = input("\nEnter your choice (1-6): ").strip()
        
        if choice == "1":
            file_path = input("Enter the path to your Excel file: ").strip()
            if os.path.exists(file_path):
                if any(file_path.lower().endswith(ext) for ext in config.get('supported_extensions')):
                    process_single_file(file_path)
                else:
                    print("❌ Please provide an Excel file (.xlsx, .xlsm, .xltx, .xltm)")
            else:
                print("❌ File not found!")
        
        elif choice == "2":
            folder_path = input("Enter the folder path: ").strip()
            batch_process_folder(folder_path)
        
        elif choice == "3":
            show_configuration()
        
        elif choice == "4":
            update_configuration()
        
        elif choice == "5":
            view_log_file()
        
        elif choice == "6":
            logging.info("=== Excel Hyperlink Bot Stopped ===")
            print("👋 Thank you for using the Enhanced Excel Bot!")
            break
        
        else:
            print("❌ Invalid choice!")

# ==================== START THE APPLICATION ====================
if __name__ == "__main__":
    main()